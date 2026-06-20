import os
import sys
import json
import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font

def parse_excel_date(val):
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, (int, float)):
        # Excel date is days since 1899-12-30 (due to 1900 leap year bug)
        try:
            base_date = datetime.datetime(1899, 12, 30)
            delta = datetime.timedelta(days=val)
            return (base_date + delta).strftime('%Y-%m-%d')
        except Exception:
            return str(val)
    if isinstance(val, str):
        val_strip = val.strip()
        if val_strip.isdigit():
            try:
                base_date = datetime.datetime(1899, 12, 30)
                delta = datetime.timedelta(days=int(val_strip))
                return (base_date + delta).strftime('%Y-%m-%d')
            except Exception:
                pass
        # If it's already in format YYYY-MM-DD
        if len(val_strip) >= 10 and val_strip[4] == '-' and val_strip[7] == '-':
            return val_strip[:10]
    return str(val)

def cmd_parse(filepath):
    if not os.path.exists(filepath):
        print(json.dumps({"error": f"File not found: {filepath}"}))
        sys.exit(1)

    wb = openpyxl.load_workbook(filepath, data_only=True)
    
    # 1. Parse 'Certificate & Survey Tracker' sheet
    sheet = wb['Certificate & Survey Tracker']
    rows = list(sheet.iter_rows(values_only=True))
    
    # Vessel header info
    vessel_info = {
        "name": rows[1][1],
        "imo_number": rows[1][3],
        "flag": rows[1][5],
        "report_date": parse_excel_date(rows[1][7]),
        "company": rows[2][1],
        "year_built": rows[2][3],
        "date_of_build": parse_excel_date(rows[2][5]),
        "class_status": rows[2][7],
        "asset_type": rows[3][1],
        "classification_society": rows[3][3],
        "overall_status": rows[3][5],
        "dwt": rows[3][7],
        "owner": rows[4][1],
        "gross_tonnage": rows[4][3],
        "port_of_registry": rows[4][5],
        "call_sign": rows[4][7],
    }

    certificates = []
    
    # Primary Certificates: Rows 8 to 22 (1-based row 9 to 23)
    for i in range(8, 23):
        if i >= len(rows):
            break
        row = rows[i]
        if row[0] is None or str(row[0]).strip() == "":
            continue
        
        name = str(row[0]).strip()
        org = str(row[1]).strip() if row[1] else None
        issue_date = parse_excel_date(row[2])
        expiry_date = parse_excel_date(row[3])
        due_date = parse_excel_date(row[4])
        window = str(row[5]).strip() if row[5] else None
        alarm = str(row[6]).strip() if row[6] else None
        remarks = str(row[7]).strip() if row[7] else None
        
        # Categorize
        name_lower = name.lower()
        if "tonnage" in name_lower or "energy efficiency" in name_lower:
            category = "Flag"
        else:
            category = "Class"
            
        certificates.append({
            "excel_row": i + 1,
            "name": name,
            "category": category,
            "organization": org,
            "issuing_date": issue_date,
            "expiration_date": expiry_date,
            "due_date": due_date,
            "window": window,
            "alarm_status": alarm,
            "remarks": remarks
        })

    # Periodical Surveys: Rows 26 to 61 (1-based row 27 to 62)
    for i in range(26, len(rows)):
        if i >= len(rows):
            break
        row = rows[i]
        if row[0] is None or str(row[0]).strip() == "" or "PERIODICAL SURVEY" in str(row[0]):
            continue
            
        name = str(row[0]).strip()
        org = str(row[1]).strip() if row[1] else None
        issue_date = parse_excel_date(row[2])
        expiry_date = parse_excel_date(row[3])
        due_date = parse_excel_date(row[4])
        window = str(row[5]).strip() if row[5] else None
        alarm = str(row[6]).strip() if row[6] else None
        remarks = str(row[7]).strip() if row[7] else None
        
        certificates.append({
            "excel_row": i + 1,
            "name": name,
            "category": "Class", # Periodical surveys from LR
            "organization": org,
            "issuing_date": issue_date,
            "expiration_date": expiry_date,
            "due_date": due_date,
            "window": window,
            "alarm_status": alarm,
            "remarks": remarks
        })

    # 2. Parse 'Actionable Items' sheet
    actionable_items = []
    if 'Actionable Items' in wb.sheetnames:
        action_sheet = wb['Actionable Items']
        act_rows = list(action_sheet.iter_rows(values_only=True))
        # Header is in row 2 (0-indexed 2, which is Row 3 in Excel)
        # Rows start at index 3 (Row 4 in Excel)
        for i in range(3, len(act_rows)):
            row = act_rows[i]
            if len(row) < 6 or row[0] is None or str(row[0]).strip() == "" or "Advice to Owners" in str(row[0]) or str(row[0]).startswith('ID'):
                continue
            
            actionable_items.append({
                "item_id": str(row[0]).strip(),
                "imposed_date": parse_excel_date(row[1]),
                "category": str(row[2]).strip() if row[2] else None,
                "report_number": str(row[3]).strip() if row[3] else None,
                "due_date": parse_excel_date(row[4]),
                "description": str(row[5]).strip() if row[5] else ""
            })

    output = {
        "vessel": vessel_info,
        "certificates": certificates,
        "actionable_items": actionable_items
    }
    print(json.dumps(output, indent=2))

def cmd_format(template_path, output_path, data_json_path):
    if not os.path.exists(template_path):
        print(f"Error: Template not found: {template_path}")
        sys.exit(1)
    if not os.path.exists(data_json_path):
        print(f"Error: JSON data file not found: {data_json_path}")
        sys.exit(1)

    with open(data_json_path, 'r') as f:
        data = json.load(f)

    # Load workbook preserving formulas
    wb = openpyxl.load_workbook(template_path, data_only=False)
    sheet = wb['Certificate & Survey Tracker']

    vessel = data.get('vessel', {})
    certificates = data.get('certificates', [])
    emails = data.get('emails', [])

    # Write Vessel Header Info
    sheet['B2'] = vessel.get('name', '')
    sheet['D2'] = vessel.get('imo_number', '')
    sheet['F2'] = vessel.get('flag', '')
    if vessel.get('report_date'):
        sheet['H2'] = datetime.datetime.strptime(vessel['report_date'], '%Y-%m-%d')

    sheet['B3'] = vessel.get('company', '')
    sheet['D3'] = vessel.get('year_built', '')
    if vessel.get('date_of_build'):
        sheet['F3'] = datetime.datetime.strptime(vessel['date_of_build'], '%Y-%m-%d')
    sheet['H3'] = vessel.get('class_status', '')

    sheet['B4'] = vessel.get('asset_type', '')
    sheet['D4'] = vessel.get('classification_society', '')
    sheet['F4'] = vessel.get('overall_status', '')
    sheet['H4'] = vessel.get('dwt', '')

    sheet['B5'] = vessel.get('owner', '')
    sheet['D5'] = vessel.get('gross_tonnage', '')
    sheet['F5'] = vessel.get('port_of_registry', '')
    sheet['H5'] = vessel.get('call_sign', '')

    # Write emails in row 6 (cells B6, C6, D6)
    for idx, email in enumerate(emails[:3]):
        col = ['B', 'C', 'D'][idx]
        sheet[f'{col}6'] = email

    # Style fills and fonts
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006", bold=True)
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    yellow_font = Font(color="9C6500", bold=True)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(color="006100", bold=True)
    gray_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
    gray_font = Font(color="595959")

    # Update certificates in Excel sheet
    for cert in certificates:
        row = cert.get('excel_row')
        if not row:
            continue
        
        # Write Dates (convert back to datetime objects for openpyxl)
        def write_date_cell(col_name, date_str):
            cell = sheet[f'{col_name}{row}']
            if date_str:
                try:
                    cell.value = datetime.datetime.strptime(date_str, '%Y-%m-%d')
                    cell.number_format = 'yyyy-mm-dd'
                except Exception:
                    cell.value = date_str
            else:
                cell.value = None

        write_date_cell('C', cert.get('issuing_date'))
        write_date_cell('D', cert.get('expiration_date'))
        write_date_cell('E', cert.get('due_date'))

        if cert.get('window') is not None:
            sheet[f'F{row}'] = cert.get('window')
        if cert.get('remarks') is not None:
            sheet[f'H{row}'] = cert.get('remarks')

        # Overwrite formula in Column G
        # G row formula
        formula = (
            f'=IF($E{row}="","N/A",'
            f'IF($E{row}<TODAY(),"OVERDUE / IMMEDIATE",'
            f'IF($E{row}<=TODAY()+30,"RED - <1 MONTH",'
            f'IF($E{row}<=TODAY()+90,"YELLOW - 1 TO 3 MONTHS",'
            f'IF($E{row}<=TODAY()+180,"GREEN - 3 TO 6 MONTHS","MONITOR >6 MONTHS")))))'
        )
        sheet[f'G{row}'] = formula

        # Apply cell-level fill colors directly for static visual compliance
        alarm_status = cert.get('alarm_status', '').upper()
        cell_g = sheet[f'G{row}']
        if 'RED' in alarm_status or 'OVERDUE' in alarm_status or 'IMMEDIATE' in alarm_status:
            cell_g.fill = red_fill
            cell_g.font = red_font
        elif 'YELLOW' in alarm_status:
            cell_g.fill = yellow_fill
            cell_g.font = yellow_font
        elif 'GREEN' in alarm_status:
            cell_g.fill = green_fill
            cell_g.font = green_font
        elif 'N/A' in alarm_status:
            cell_g.fill = gray_fill
            cell_g.font = gray_font
        else:
            # Clear formatting/use gray
            cell_g.fill = gray_fill
            cell_g.font = gray_font

    # Actionable items sheet
    if 'Actionable Items' in wb.sheetnames and 'actionable_items' in data:
        act_sheet = wb['Actionable Items']
        # Clear existing actionable items rows starting from row 4
        # We can rewrite them
        # Let's delete from row 4 to 12
        for r_idx in range(4, 15):
            for c_idx in range(1, 7):
                act_sheet.cell(row=r_idx, column=c_idx).value = None
        
        # Write new list
        for idx, act in enumerate(data['actionable_items']):
            row_idx = 4 + idx
            act_sheet.cell(row=row_idx, column=1, value=act.get('item_id', ''))
            
            # Imposed date
            imp_date = act.get('imposed_date')
            if imp_date:
                cell_imp = act_sheet.cell(row=row_idx, column=2)
                try:
                    cell_imp.value = datetime.datetime.strptime(imp_date, '%Y-%m-%d')
                    cell_imp.number_format = 'yyyy-mm-dd'
                except Exception:
                    cell_imp.value = imp_date
                    
            act_sheet.cell(row=row_idx, column=3, value=act.get('category', ''))
            act_sheet.cell(row=row_idx, column=4, value=act.get('report_number', ''))
            
            # Due date
            due_dt = act.get('due_date')
            if due_dt:
                cell_due = act_sheet.cell(row=row_idx, column=5)
                try:
                    cell_due.value = datetime.datetime.strptime(due_dt, '%Y-%m-%d')
                    cell_due.number_format = 'yyyy-mm-dd'
                except Exception:
                    cell_due.value = due_dt
                    
            act_sheet.cell(row=row_idx, column=6, value=act.get('description', ''))

    wb.save(output_path)
    print(json.dumps({"success": True, "output_file": output_path}))

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: python excel_handler.py parse <filepath>")
        print("Usage: python excel_handler.py format <template_path> <output_path> <json_data_path>")
        sys.exit(1)
        
    cmd = sys.argv[1]
    if cmd == 'parse':
        cmd_parse(sys.argv[2])
    elif cmd == 'format':
        if len(sys.argv) < 5:
            print("Missing parameters for format command.")
            sys.exit(1)
        cmd_format(sys.argv[2], sys.argv[3], sys.argv[4])
    else:
        print(f"Unknown command: {cmd}")
        sys.exit(1)
